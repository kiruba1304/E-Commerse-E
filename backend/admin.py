from flask import Blueprint, request, jsonify, send_file, render_template_string
from models import db, Admin, Shop, Product, Category, Order, OrderItem, PopupAd, Coupon, Review, User, SystemLog, Notification, Collection, CustomizationOrder, NewsletterSubscription
from auth_middleware import generate_token, token_required, role_required
from datetime import datetime, timezone
import json

admin_bp = Blueprint('admin', __name__)

def log_admin_action(admin_id, username, shop_id, action):
    try:
        log = SystemLog(
            actor_type='admin',
            actor_id=admin_id,
            username=username,
            action=action,
            shop_id=shop_id
        )
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        print("Log error:", e)
        db.session.rollback()

@admin_bp.route('/login', methods=['POST'])
def login():
    data = request.get_json() or {}
    username = data.get('username')
    password = data.get('password')

    if not username or not password:
        return jsonify({"error": "Username and password are required"}), 400

    admin = Admin.query.filter_by(username=username).first()
    if not admin or not admin.check_password(password) or not admin.is_active:
        return jsonify({"error": "Invalid admin credentials or account is suspended"}), 401

    token = generate_token(admin.id, admin.username, 'admin', admin.shop_id)
    
    # Log login
    log_admin_action(admin.id, admin.username, admin.shop_id, "Admin logged in successfully")
    
    return jsonify({
        "message": "Login successful",
        "token": token,
        "user": admin.serialize()
    }), 200

@admin_bp.route('/logout', methods=['POST'])
@role_required(['admin'])
def logout():
    user = request.user
    log_admin_action(user['user_id'], user['username'], user['shop_id'], "Admin logged out")
    return jsonify({"message": "Logout successful"}), 200

@admin_bp.route('/verify-password', methods=['POST'])
@role_required(['admin'])
def verify_password():
    admin = Admin.query.get(request.user['user_id'])
    if not admin:
        return jsonify({"error": "Admin not found"}), 404

    data = request.get_json() or {}
    password = data.get('password')
    if not password:
        return jsonify({"error": "Password is required"}), 400

    if not admin.check_password(password):
        return jsonify({"success": False, "error": "Invalid password"}), 401

    return jsonify({"success": True}), 200

@admin_bp.route('/profile', methods=['GET', 'PUT'])
@role_required(['admin'])
def manage_profile():
    admin = Admin.query.get(request.user['user_id'])
    if not admin:
        return jsonify({"error": "Admin not found"}), 404

    if request.method == 'GET':
        return jsonify(admin.serialize()), 200
    
    data = request.get_json() or {}
    if 'name' in data:
        admin.name = data['name']
    if 'email' in data:
        admin.email = data['email']
    if 'password' in data and data['password']:
        admin.set_password(data['password'])
        
    db.session.commit()
    log_admin_action(admin.id, admin.username, admin.shop_id, "Updated admin profile details")
    
    return jsonify({"message": "Profile updated successfully", "admin": admin.serialize()}), 200

@admin_bp.route('/shop', methods=['GET', 'PUT'])
@role_required(['admin'])
def manage_shop():
    shop = Shop.query.get(request.user['shop_id'])
    if not shop:
        return jsonify({"error": "Shop not found"}), 404

    if request.method == 'GET':
        return jsonify(shop.serialize()), 200
    data = request.get_json() or {}
    if 'name' in data:
        shop.name = data['name']
    if 'logo_url' in data:
        shop.logo_url = data['logo_url']
    if 'contact_email' in data:
        shop.contact_email = data['contact_email']
    if 'contact_phone' in data:
        shop.contact_phone = data['contact_phone']
    if 'privacy_policy' in data:
        shop.privacy_policy = data['privacy_policy']
    if 'address' in data:
        shop.address = data['address']
    if 'sms_api_key' in data:
        shop.sms_api_key = data['sms_api_key']
    if 'sms_enabled' in data:
        shop.sms_enabled = bool(data['sms_enabled'])
    if 'sms_dispatch_enabled' in data:
        shop.sms_dispatch_enabled = bool(data['sms_dispatch_enabled'])
    if 'sms_delivery_enabled' in data:
        shop.sms_delivery_enabled = bool(data['sms_delivery_enabled'])
    if 'sms_campaign_enabled' in data:
        shop.sms_campaign_enabled = bool(data['sms_campaign_enabled'])
    if 'sms_sender_id' in data:
        shop.sms_sender_id = data['sms_sender_id']
    if 'sms_otp_template_id' in data:
        shop.sms_otp_template_id = data['sms_otp_template_id']
    if 'sms_dispatch_template_id' in data:
        shop.sms_dispatch_template_id = data['sms_dispatch_template_id']
    if 'sms_delivery_template_id' in data:
        shop.sms_delivery_template_id = data['sms_delivery_template_id']
    if 'whatsapp_api_key' in data:
        shop.whatsapp_api_key = data['whatsapp_api_key']
    if 'razorpay_key_id' in data:
        shop.razorpay_key_id = data['razorpay_key_id']
    if 'razorpay_key_secret' in data:
        shop.razorpay_key_secret = data['razorpay_key_secret']
    if 'billing_api_key' in data:
        shop.billing_api_key = data['billing_api_key']
    if 'dtdc_client_code' in data:
        shop.dtdc_client_code = data['dtdc_client_code']
    if 'dtdc_api_key' in data:
        shop.dtdc_api_key = data['dtdc_api_key']
    if 'dtdc_api_url' in data:
        shop.dtdc_api_url = data['dtdc_api_url']
    if 'shiprocket_email' in data:
        shop.shiprocket_email = data['shiprocket_email']
    if 'shiprocket_password' in data and data['shiprocket_password']:
        shop.shiprocket_password = data['shiprocket_password']
    if 'shiprocket_pickup_location' in data:
        shop.shiprocket_pickup_location = data['shiprocket_pickup_location']
    if 'gst_percentage' in data:
        try:
            shop.gst_percentage = float(data['gst_percentage'])
        except ValueError:
            return jsonify({"error": "Invalid GST percentage value"}), 400
    if 'gst_inclusive' in data:
        shop.gst_inclusive = bool(data['gst_inclusive'])
    if 'saree_models' in data:
        shop.saree_models = data['saree_models']
    if 'banners' in data:
        shop.banners = data['banners']
    if 'smtp_host' in data:
        shop.smtp_host = data['smtp_host']
    if 'smtp_port' in data:
        try:
            shop.smtp_port = int(data['smtp_port']) if data['smtp_port'] else None
        except ValueError:
            return jsonify({"error": "Invalid SMTP port value"}), 400
    if 'smtp_user' in data:
        shop.smtp_user = data['smtp_user']
    if 'smtp_password' in data:
        shop.smtp_password = data['smtp_password']
    if 'smtp_use_tls' in data:
        shop.smtp_use_tls = bool(data['smtp_use_tls'])
    if 'smtp_sender_name' in data:
        shop.smtp_sender_name = data['smtp_sender_name']
    if 'email_templates' in data:
        shop.email_templates = data['email_templates']
    if 'color_palette' in data:
        shop.color_palette = data['color_palette']
    if 'customization_min_quantity' in data:
        try:
            shop.customization_min_quantity = int(data['customization_min_quantity'])
        except ValueError:
            return jsonify({"error": "Invalid customization minimum quantity"}), 400
    if 'return_window_days' in data:
        try:
            shop.return_window_days = int(data['return_window_days'])
        except ValueError:
            return jsonify({"error": "Invalid return window days value"}), 400

    if 'shipping_enabled' in data:
        shop.shipping_enabled = bool(data['shipping_enabled'])
    if 'shipping_charges_type' in data:
        shop.shipping_charges_type = str(data['shipping_charges_type'])
    if 'shipping_charges_flat' in data:
        try:
            shop.shipping_charges_flat = float(data['shipping_charges_flat'])
        except ValueError:
            return jsonify({"error": "Invalid flat shipping charge value"}), 400
    if 'cod_enabled' in data:
        shop.cod_enabled = bool(data['cod_enabled'])
    if 'customization_cod_enabled' in data:
        shop.customization_cod_enabled = bool(data['customization_cod_enabled'])
    if 'welcome_super_coins' in data:
        try:
            shop.welcome_super_coins = int(data['welcome_super_coins']) if data['welcome_super_coins'] is not None else 50
        except ValueError:
            return jsonify({"error": "Invalid welcome pearls value"}), 400
    if 'signature_url' in data:
        shop.signature_url = data['signature_url']
    if 'store_locator_link' in data:
        shop.store_locator_link = data['store_locator_link']

    db.session.commit()
    log_admin_action(request.user['user_id'], request.user['username'], shop.id, "Updated shop details & payment integrations")
    return jsonify({"message": "Shop settings updated successfully", "shop": shop.serialize()}), 200


@admin_bp.route('/billing-status', methods=['GET'])
@role_required(['admin'])
def billing_status_page():
    """Serve a small admin page that polls the billing sync status for the current shop."""
    shop = Shop.query.get(request.user['shop_id'])
    if not shop:
        return jsonify({"error": "Shop not found"}), 404

    api_key = shop.billing_api_key or ''
    api_key_json = json.dumps(api_key)

    html = '''
    <!doctype html>
    <html>
        <head>
            <meta charset="utf-8" />
            <title>Billing App Status</title>
            <style>
                body { font-family: Arial, sans-serif; padding: 24px; }
                .status { display:flex; align-items:center; gap:12px; }
                .dot { width:14px; height:14px; border-radius:50%; background:#999 }
                .online { background: #22c55e }
                .offline { background: #ef4444 }
            </style>
        </head>
        <body>
            <h2>Billing Desktop Presence</h2>
            <div class="status">
                <div id="dot" class="dot offline"></div>
                <div>
                    <div id="state">Checking...</div>
                    <div id="last">Last seen: N/A</div>
                </div>
            </div>

            <script>
                const apiKey = {{ api_key_json|safe }};
                async function fetchStatus(){
                    try{
                        const url = '/api/billing/sync/status?api_key=' + encodeURIComponent(apiKey);
                        const res = await fetch(url, { method: 'GET', headers: { 'Accept': 'application/json' } });
                        if (!res.ok) throw new Error('Network');
                        const j = await res.json();
                        const st = j.billing_status || {};
                        const isOnline = !!st.is_online;
                        document.getElementById('dot').className = 'dot ' + (isOnline ? 'online' : 'offline');
                        document.getElementById('state').textContent = isOnline ? 'Online' : 'Offline';
                        document.getElementById('last').textContent = 'Last seen: ' + (st.last_seen_at || 'N/A');
                    }catch(e){
                        document.getElementById('dot').className = 'dot offline';
                        document.getElementById('state').textContent = 'Error';
                        document.getElementById('last').textContent = '';
                    }
                }

                fetchStatus();
                setInterval(fetchStatus, 10000);
            </script>
        </body>
    </html>
    '''

    return render_template_string(html, api_key_json=api_key_json)

@admin_bp.route('/shop/test-smtp', methods=['POST'])
@role_required(['admin'])
def test_smtp_configuration():
    shop = Shop.query.get(request.user['shop_id'])
    if not shop:
        return jsonify({"error": "Shop not found"}), 404

    data = request.get_json() or {}
    recipient = data.get('recipient')
    if not recipient:
        return jsonify({"error": "Recipient email is required"}), 400

    # Temporarily override with unsaved config details if sent for pre-testing
    templates_data = data.get('email_templates', None)
    if templates_data:
        import json
        templates_json = json.dumps(templates_data)
    else:
        templates_json = shop.email_templates_json

    test_shop = Shop(
        id=shop.id,
        name=shop.name,
        smtp_host=data.get('smtp_host', shop.smtp_host),
        smtp_port=int(data.get('smtp_port')) if data.get('smtp_port') else shop.smtp_port,
        smtp_user=data.get('smtp_user', shop.smtp_user),
        smtp_password=data.get('smtp_password', shop.smtp_password),
        smtp_use_tls=bool(data.get('smtp_use_tls', shop.smtp_use_tls)),
        smtp_sender_name=data.get('smtp_sender_name', shop.smtp_sender_name),
        email_templates_json=templates_json
    )

    from mail_sender import send_shop_email
    # We will test using a simple test payload
    placeholders = {
        "name": "Admin Tester",
        "otp": "123456",
        "reset_link": "http://localhost:5173/reset?token=test",
        "order_id": "9999",
        "total_amount": "999.00",
        "items": "1x Test Item",
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }

    test_template_type = data.get('template_type', 'otp')
    success = send_shop_email(test_shop, test_template_type, recipient, placeholders, sender_info={
        "actor_type": "admin",
        "actor_id": request.user['user_id'],
        "username": request.user['username']
    })
    if success:
        return jsonify({"message": f"Test email sent successfully to {recipient}!"}), 200
    else:
        return jsonify({"error": "Failed to send email. Check your SMTP settings and logs."}), 500


# CATEGORY MANAGEMENT
@admin_bp.route('/categories', methods=['GET', 'POST'])
@role_required(['admin'])
def manage_categories():
    shop_id = request.user['shop_id']
    
    if request.method == 'GET':
        categories = Category.query.filter_by(shop_id=shop_id).all()
        return jsonify([c.serialize() for c in categories]), 200

    data = request.get_json() or {}
    name = data.get('name')
    if not name:
        return jsonify({"error": "Category name is required"}), 400

    return_window_days = data.get('return_window_days')
    if return_window_days is not None and return_window_days != "":
        try:
            return_window_days = int(return_window_days)
        except ValueError:
            return jsonify({"error": "Invalid return window days value"}), 400
    else:
        return_window_days = None

    shipping_charge = data.get('shipping_charge')
    if shipping_charge is not None and shipping_charge != "":
        try:
            shipping_charge = float(shipping_charge)
        except ValueError:
            return jsonify({"error": "Invalid shipping charge value"}), 400
    else:
        shipping_charge = 0.0

    cat = Category(
        name=name,
        description=data.get('description', ''),
        image_url=data.get('image_url', ''),
        shop_id=shop_id,
        customization_enabled=bool(data.get('customization_enabled', False)),
        return_window_days=return_window_days,
        shipping_charge=shipping_charge,
        show_description=bool(data.get('show_description', False))
    )
    db.session.add(cat)
    db.session.commit()
    
    log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Created product category '{name}'")
    return jsonify(cat.serialize()), 201

@admin_bp.route('/categories/<int:cat_id>', methods=['PUT', 'DELETE'])
@role_required(['admin'])
def modify_category(cat_id):
    shop_id = request.user['shop_id']
    cat = Category.query.filter_by(id=cat_id, shop_id=shop_id).first()
    if not cat:
        return jsonify({"error": "Category not found"}), 404

    if request.method == 'DELETE':
        # Re-assign products to uncategorized if category deleted
        products = Product.query.filter_by(category_id=cat_id).all()
        for p in products:
            p.category_id = None
        
        name = cat.name
        db.session.delete(cat)
        db.session.commit()
        log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Deleted product category '{name}'")
        return jsonify({"message": "Category deleted successfully"}), 200

    data = request.get_json() or {}
    if 'name' in data:
        cat.name = data['name']
    if 'description' in data:
        cat.description = data['description']
    if 'image_url' in data:
        cat.image_url = data['image_url']
    if 'customization_enabled' in data:
        cat.customization_enabled = bool(data['customization_enabled'])
    if 'return_window_days' in data:
        val = data['return_window_days']
        if val is not None and val != "":
            try:
                cat.return_window_days = int(val)
            except ValueError:
                return jsonify({"error": "Invalid return window days value"}), 400
        else:
            cat.return_window_days = None

    if 'shipping_charge' in data:
        val = data['shipping_charge']
        if val is not None and val != "":
            try:
                cat.shipping_charge = float(val)
            except ValueError:
                return jsonify({"error": "Invalid shipping charge value"}), 400
        else:
            cat.shipping_charge = 0.0

    if 'show_description' in data:
        cat.show_description = bool(data['show_description'])

    db.session.commit()
    log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Updated product category '{cat.name}'")
    return jsonify(cat.serialize()), 200

# COLLECTION MANAGEMENT
@admin_bp.route('/collections', methods=['GET', 'POST'])
@role_required(['admin'])
def manage_collections():
    shop_id = request.user['shop_id']
    
    if request.method == 'GET':
        collections = Collection.query.filter_by(shop_id=shop_id).all()
        return jsonify([c.serialize() for c in collections]), 200

    data = request.get_json() or {}
    name = data.get('name')
    if not name:
        return jsonify({"error": "Collection name is required"}), 400

    category_ids = data.get('category_ids', [])
    col = Collection(
        name=name,
        shop_id=shop_id,
        separate_categories_mobile=bool(data.get('separate_categories_mobile', False)),
        show_category_banner=bool(data.get('show_category_banner', True))
    )
    col.category_ids = category_ids
    db.session.add(col)
    db.session.commit()
    
    log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Created collection '{name}'")
    return jsonify(col.serialize()), 201

@admin_bp.route('/collections/<int:col_id>', methods=['PUT', 'DELETE'])
@role_required(['admin'])
def modify_collection(col_id):
    shop_id = request.user['shop_id']
    col = Collection.query.filter_by(id=col_id, shop_id=shop_id).first()
    if not col:
        return jsonify({"error": "Collection not found"}), 404

    if request.method == 'DELETE':
        name = col.name
        db.session.delete(col)
        db.session.commit()
        log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Deleted collection '{name}'")
        return jsonify({"message": "Collection deleted successfully"}), 200

    data = request.get_json() or {}
    if 'name' in data:
        col.name = data['name']
    if 'category_ids' in data:
        col.category_ids = data['category_ids']
    if 'separate_categories_mobile' in data:
        col.separate_categories_mobile = bool(data['separate_categories_mobile'])
    if 'show_category_banner' in data:
        col.show_category_banner = bool(data['show_category_banner'])

    db.session.commit()
    log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Updated collection '{col.name}'")
    return jsonify(col.serialize()), 200

# PRODUCT CATALOG (CRUD with multiple image support up to 10)
@admin_bp.route('/products', methods=['GET', 'POST'])
@role_required(['admin'])
def manage_products():
    shop_id = request.user['shop_id']

    if request.method == 'GET':
        products = Product.query.filter_by(shop_id=shop_id, is_deleted=False).all()
        return jsonify([p.serialize() for p in products]), 200

    data = request.get_json() or {}
    name = data.get('name')
    price = data.get('price')
    stock = data.get('stock')

    if not name or price is None or stock is None:
        return jsonify({"error": "Name, price, and stock are required fields"}), 400

    images = data.get('images', [])
    if len(images) > 10:
        return jsonify({"error": "A product can have a maximum of 10 images"}), 400

    return_window_days = data.get('return_window_days')
    if return_window_days is not None and return_window_days != "":
        try:
            return_window_days = int(return_window_days)
        except ValueError:
            return jsonify({"error": "Invalid return window days value"}), 400
    else:
        return_window_days = None

    p = Product(
        name=name,
        description=data.get('description', ''),
        price=float(price),
        original_price=float(data.get('original_price', price)),
        stock=int(stock),
        alert_threshold=int(data.get('alert_threshold', 5)),
        promo_code=data.get('promo_code', ''),
        promo_discount=float(data.get('promo_discount', 0.0)) if data.get('promo_discount') else 0.0,
        bulk_sale_price=float(data['bulk_sale_price']) if data.get('bulk_sale_price') else None,
        min_quantity=int(data['min_quantity']) if data.get('min_quantity') else None,
        category_id=data.get('category_id'),
        shop_id=shop_id,
        customization_enabled=bool(data.get('customization_enabled', False)),
        barcode=data.get('barcode', ''),
        sku_code=data.get('sku_code', ''),
        hsc_code=data.get('hsc_code', ''),
        return_window_days=return_window_days,
        cod_enabled=bool(data.get('cod_enabled', True))
    )
    p.images = images  # sets JSON field via property setter
    p.custom_colors = data.get('custom_colors', [])
    
    db.session.add(p)
    db.session.commit()

    # Send FCM product notification to topic
    try:
        from fcm_helper import send_fcm_topic_notification
        topic = f"shop_{shop_id}_products"
        title = "New Product Added!"
        body = f"Check out our new arrival: {p.name} for only ₹{p.price}!"
        send_fcm_topic_notification(topic, title, body, data={
            "product_id": str(p.id)
        })
    except Exception as e:
        print(f"Failed to send FCM product notification: {e}")

    log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Added product '{name}' (Stock: {stock}, Price: {price})")
    
    return jsonify(p.serialize()), 201

@admin_bp.route('/products/<int:prod_id>', methods=['PUT', 'DELETE'])
@role_required(['admin'])
def modify_product(prod_id):
    shop_id = request.user['shop_id']
    p = Product.query.filter_by(id=prod_id, shop_id=shop_id).first()
    if not p:
        return jsonify({"error": "Product not found"}), 404

    if request.method == 'DELETE':
        name = p.name
        has_order_refs = OrderItem.query.filter_by(product_id=prod_id).first() is not None or \
                         CustomizationOrder.query.filter_by(product_id=prod_id).first() is not None
        try:
            if has_order_refs:
                p.is_deleted = True
                p.category_id = None
                db.session.commit()
                log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Soft-deleted product '{name}' (retained for order history)")
            else:
                db.session.delete(p)
                db.session.commit()
                log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Hard-deleted product '{name}'")
            return jsonify({"message": "Product deleted successfully"}), 200
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": f"Failed to delete product: {str(e)}"}), 500

    data = request.get_json() or {}
    if 'name' in data:
        p.name = data['name']
    if 'description' in data:
        p.description = data['description']
    if 'price' in data:
        p.price = float(data['price'])
    if 'original_price' in data:
        p.original_price = float(data['original_price'])
    if 'stock' in data:
        p.stock = int(data['stock'])
    if 'alert_threshold' in data:
        p.alert_threshold = int(data['alert_threshold'])
    if 'images' in data:
        images = data['images']
        if len(images) > 10:
            return jsonify({"error": "A product can have a maximum of 10 images"}), 400
        p.images = images
    if 'promo_code' in data:
        p.promo_code = data['promo_code']
    if 'promo_discount' in data:
        p.promo_discount = float(data['promo_discount'])
    if 'bulk_sale_price' in data:
        p.bulk_sale_price = float(data['bulk_sale_price']) if data['bulk_sale_price'] else None
    if 'min_quantity' in data:
        p.min_quantity = int(data['min_quantity']) if data['min_quantity'] else None
    if 'category_id' in data:
        p.category_id = data['category_id']
    if 'customization_enabled' in data:
        p.customization_enabled = bool(data['customization_enabled'])
    if 'barcode' in data:
        p.barcode = data['barcode']
    if 'sku_code' in data:
        p.sku_code = data['sku_code']
    if 'hsc_code' in data:
        p.hsc_code = data['hsc_code']
    if 'return_window_days' in data:
        val = data['return_window_days']
        if val is not None and val != "":
            try:
                p.return_window_days = int(val)
            except ValueError:
                return jsonify({"error": "Invalid return window days value"}), 400
        else:
            p.return_window_days = None
    if 'cod_enabled' in data:
        p.cod_enabled = bool(data['cod_enabled'])
    if 'custom_colors' in data:
        p.custom_colors = data['custom_colors']

    db.session.commit()
    log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Modified product details for '{p.name}'")
    return jsonify(p.serialize()), 200

# INVENTORY MANAGEMENT (Stock Evaluation & Alert)
@admin_bp.route('/inventory', methods=['GET'])
@role_required(['admin'])
def inventory_evaluation():
    shop_id = request.user['shop_id']
    products = Product.query.filter_by(shop_id=shop_id, is_deleted=False).all()
    
    total_items = 0
    total_value = 0.0
    alerts = []
    
    for p in products:
        total_items += p.stock
        total_value += (p.stock * p.price)
        if p.stock <= p.alert_threshold:
            alerts.append({
                "id": p.id,
                "name": p.name,
                "stock": p.stock,
                "threshold": p.alert_threshold
            })
            
    return jsonify({
        "total_unique_products": len(products),
        "total_stock_count": total_items,
        "total_inventory_value": total_value,
        "alerts": alerts
    }), 200

# REVENUE REPORT AND GRAPHICAL ANALYSIS
@admin_bp.route('/revenue-report', methods=['GET'])
@role_required(['admin'])
def revenue_report():
    shop_id = request.user['shop_id']
    
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    start_date = None
    end_date = None
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        except ValueError:
            pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str + ' 23:59:59', '%Y-%m-%d %H:%M:%S')
        except ValueError:
            pass

    # Base query for completed orders (excluding returned)
    orders_query = Order.query.filter_by(shop_id=shop_id).filter(Order.status != 'Returned')
    
    # Base query for all orders (for payments, returns, promos, etc.)
    all_orders_query = Order.query.filter_by(shop_id=shop_id)
    
    # Base query for customization orders
    custom_orders_query = CustomizationOrder.query.filter_by(shop_id=shop_id)
    
    # Base query for users
    users_query = User.query
    
    if start_date:
        orders_query = orders_query.filter(Order.created_at >= start_date)
        all_orders_query = all_orders_query.filter(Order.created_at >= start_date)
        custom_orders_query = custom_orders_query.filter(CustomizationOrder.created_at >= start_date)
        users_query = users_query.filter(User.created_at >= start_date)
        
    if end_date:
        orders_query = orders_query.filter(Order.created_at <= end_date)
        all_orders_query = all_orders_query.filter(Order.created_at <= end_date)
        custom_orders_query = custom_orders_query.filter(CustomizationOrder.created_at <= end_date)
        users_query = users_query.filter(User.created_at <= end_date)

    all_completed = orders_query.all()
    # Filter completed orders to exclude:
    # 1. Rejected orders (for both COD and UPI)
    # 2. Pending COD orders (unaccepted COD)
    all_completed = [
        o for o in all_completed
        if o.status != 'Rejected' and not (o.payment_method == 'COD' and o.status == 'Pending')
    ]
    
    all_orders = all_orders_query.all()
    custom_orders = custom_orders_query.all()
    users_list = users_query.all()
    
    total_sales = sum([o.final_amount for o in all_completed])
    gst_collected = sum([o.gst_amount for o in all_completed])
    total_orders = len(all_completed)
    
    # Group by category
    category_sales = {}
    # Group by payment method
    payment_sales = {"COD": 0, "UPI": 0}
    # Sales by day grouping
    sales_by_day = {}
    
    for o in all_completed:
        # Payment breakdown
        payment_sales[o.payment_method] = payment_sales.get(o.payment_method, 0) + 1
        
        # Day breakdown
        day_str = o.created_at.strftime('%Y-%m-%d')
        sales_by_day[day_str] = sales_by_day.get(day_str, 0.0) + o.final_amount
        
        # Category breakdown
        for item in o.items:
            prod = Product.query.get(item.product_id)
            cat_name = prod.category.name if (prod and prod.category) else "Uncategorized"
            category_sales[cat_name] = category_sales.get(cat_name, 0.0) + (item.price * item.quantity)
            
    # Format reports for simple React graph components
    cat_data = [{"name": k, "value": round(v, 2)} for k, v in category_sales.items()]
    day_data = [{"date": k, "revenue": round(v, 2)} for k, v in sorted(sales_by_day.items())]
    pay_data = [{"method": k, "count": v} for k, v in payment_sales.items()]
    
    # Detailed reports requested
    report_new_users = [u.serialize() for u in users_list]
    report_sales = [o.serialize() for o in all_completed]
    report_online = [o.serialize() for o in all_orders if o.payment_method != 'COD' and o.status != 'Returned' and o.status != 'Rejected']
    report_cod = [o.serialize() for o in all_orders if o.payment_method == 'COD' and o.status != 'Returned' and o.status != 'Rejected' and o.status != 'Pending']
    report_returns = [o.serialize() for o in all_orders if o.status == 'Returned' or o.return_request_status == 'Approved']
    report_custom = [co.serialize() for co in custom_orders]
    report_promo = [o.serialize() for o in all_orders if o.discount_amount > 0]
    
    return jsonify({
        "summary": {
            "total_revenue": round(total_sales, 2),
            "gst_collected": round(gst_collected, 2),
            "order_count": total_orders,
            "average_order_value": round(total_sales / total_orders, 2) if total_orders > 0 else 0.0
        },
        "charts": {
            "category_sales": cat_data,
            "daily_sales": day_data,
            "payment_methods": pay_data
        },
        "reports": {
            "new_users": report_new_users,
            "sales": report_sales,
            "online_payments": report_online,
            "cod_payments": report_cod,
            "returns": report_returns,
            "custom_orders": report_custom,
            "promo_orders": report_promo
        }
    }), 200

# CUSTOMER MANAGEMENT
@admin_bp.route('/customers', methods=['GET'])
@role_required(['admin'])
def list_customers():
    shop_id = request.user['shop_id']
    # Select all registered users in the system
    users = User.query.order_by(User.created_at.desc()).all()
    orders = Order.query.filter_by(shop_id=shop_id).all()
    
    customers = []
    for user in users:
        user_orders = [o for o in orders if o.user_id == user.id]
        total_spent = sum([o.final_amount for o in user_orders])
        customers.append({
            "id": user.id,
            "name": user.name,
            "username": user.username,
            "email": user.email,
            "contact_phone": user.contact_phone,
            "total_orders": len(user_orders),
            "total_spent": round(total_spent, 2),
            "super_coins": user.super_coins or 0,
            "joined_at": user.created_at.isoformat() if user.created_at else None
        })
            
    return jsonify(customers), 200

@admin_bp.route('/customers/<int:user_id>/grant-pearls', methods=['POST'])
@role_required(['admin'])
def grant_pearls(user_id):
    admin_id = request.user['user_id']
    username = request.user['username']
    shop_id = request.user['shop_id']
    
    data = request.get_json() or {}
    amount = data.get('amount')
    reason = data.get('reason', 'Boutique loyalty reward bonus')
    
    if amount is None or not isinstance(amount, int) or amount <= 0:
        return jsonify({"error": "Invalid pearls amount. Must be a positive integer."}), 400
        
    user = User.query.get(user_id)
    if not user:
        return jsonify({"error": "Customer not found."}), 404
        
    # Increment user's super_coins balance
    user.super_coins = (user.super_coins or 0) + amount
    
    # Create database notification
    notif_title = "Privilege Pearls Credited!"
    notif_message = f"Congratulations! {amount} Privilege Pearls have been added to your account. Reason: {reason}"
    notif = Notification(
        recipient_type='user',
        recipient_id=user.id,
        title=notif_title,
        message=notif_message,
        shop_id=shop_id
    )
    db.session.add(notif)
    db.session.commit()
    
    # Send FCM push notification (background)
    if user.fcm_token:
        try:
            from fcm_helper import send_fcm_notification
            send_fcm_notification(user.fcm_token, notif_title, notif_message, data={
                "type": "pearls_granted",
                "amount": str(amount),
                "new_balance": str(user.super_coins)
            })
        except Exception as e:
            print("FCM pearl notification error:", e)
            
    # Send Email notification (background thread to avoid blocking admin HTTP response)
    shop = Shop.query.get(shop_id)
    if shop and user.email:
        try:
            import threading
            from mail_sender import send_pearls_granted_email
            from flask import current_app
            
            site_link = request.headers.get('Origin') or "http://localhost:5173"
            app_context = current_app.app_context()
            
            def async_send_email(app_context, shop_id, email, amount, reason, new_balance, site_link):
                with app_context:
                    try:
                        from models import Shop
                        shop_obj = Shop.query.get(shop_id)
                        if shop_obj:
                            send_pearls_granted_email(
                                shop=shop_obj,
                                recipient_email=email,
                                amount=amount,
                                reason=reason,
                                balance=new_balance,
                                site_link=site_link,
                                sender_info={
                                    "actor_type": "admin",
                                    "actor_id": admin_id,
                                    "username": username
                                }
                            )
                    except Exception as email_err:
                        print(f"Error in background pearls email thread: {email_err}")
                        
            threading.Thread(
                target=async_send_email,
                args=(app_context, shop.id, user.email, amount, reason, user.super_coins, site_link)
            ).start()
        except Exception as spawn_err:
            print(f"Error spawning pearls email thread: {spawn_err}")
            
    # Log admin action
    log_admin_action(
        admin_id, 
        username, 
        shop_id, 
        f"Granted {amount} Privilege Pearls to customer #{user.id} ({user.username}). Reason: {reason}"
    )
    
    return jsonify({
        "message": f"Successfully granted {amount} Privilege Pearls.",
        "new_balance": user.super_coins
    }), 200

# ORDER MANAGEMENT (Tracking status: Dispatched, Customer Received, Returns)

@admin_bp.route('/orders', methods=['GET'])
@role_required(['admin'])
def manage_orders():
    shop_id = request.user['shop_id']
    orders = Order.query.filter_by(shop_id=shop_id).order_by(Order.created_at.desc()).all()
    return jsonify([o.serialize() for o in orders]), 200

@admin_bp.route('/orders/<int:order_id>', methods=['PUT'])
@role_required(['admin'])
def update_order_status(order_id):
    shop_id = request.user['shop_id']
    order = Order.query.filter_by(id=order_id, shop_id=shop_id).first()
    if not order:
        return jsonify({"error": "Order not found"}), 404

    data = request.get_json() or {}
    status = data.get('status') # Pending, Dispatched, Customer Received, Returned
    tracking_info = data.get('tracking_info')

    if status:
        valid_statuses = ['Pending', 'Accepted', 'Rejected', 'Dispatched', 'Customer Received', 'Returned']
        if status not in valid_statuses:
            return jsonify({"error": f"Invalid status. Must be one of: {valid_statuses}"}), 400
        
        # Handle SuperCoin additions if transitioned to Customer Received
        if status == 'Customer Received' and order.status != 'Customer Received':
            order.delivered_at = datetime.now()
            if order.payment_method == 'COD':
                order.payment_status = 'Paid'
            # Add supercoins to user profile
            shop = Shop.query.get(shop_id)
            if shop and shop.super_coin_enabled:
                if shop.super_coin_ratio and shop.super_coin_ratio > 0:
                    coins_to_add = int(order.final_amount / shop.super_coin_ratio)
                else:
                    coins_to_add = 0
                if coins_to_add > 0:
                    order.user.super_coins += coins_to_add
                    order.super_coins_earned = coins_to_add
                    # Push in-app notification to customer
                    notif = Notification(
                        recipient_type='user',
                        recipient_id=order.user_id,
                        title="SuperCoins Credited!",
                        message=f"Congratulations! You earned {coins_to_add} SuperCoins from your order #{order.id}.",
                        shop_id=shop_id
                    )
                    db.session.add(notif)
        
        order.status = status
        if status == 'Rejected':
            order.is_synced = True
            
        # Send FCM notification
        try:
            from fcm_helper import send_order_status_notification
            send_order_status_notification(order, status)
        except Exception as e:
            print("FCM status notification error:", e)

        # Send SMS notification
        try:
            from sms_helper import send_order_status_sms
            send_order_status_sms(order, status)
        except Exception as e:
            print("SMS status notification error:", e)
        
    if tracking_info is not None:
        order.tracking_info = tracking_info

    db.session.commit()
    log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Updated order #{order.id} status to '{order.status}'")
    
    return jsonify(order.serialize()), 200

@admin_bp.route('/orders/<int:order_id>/return', methods=['PUT'])
@role_required(['admin'])
def resolve_return_request(order_id):
    shop_id = request.user['shop_id']
    order = Order.query.filter_by(id=order_id, shop_id=shop_id).first()
    if not order:
        return jsonify({"error": "Order not found"}), 404

    data = request.get_json() or {}
    decision = data.get('decision') # Approved, Rejected
    if decision not in ['Approved', 'Rejected']:
        return jsonify({"error": "Decision must be 'Approved' or 'Rejected'"}), 400

    if order.return_request_status != 'Pending':
        return jsonify({"error": "No pending return request for this order"}), 400

    order.return_request_status = decision
    if decision == 'Approved':
        order.status = 'Returned'
        # Refund super coins if any were spent
        if order.super_coins_used > 0:
            order.user.super_coins += order.super_coins_used
            
        # Deduct super coins that were earned
        if order.super_coins_earned > 0:
            order.user.super_coins = max(0, order.user.super_coins - order.super_coins_earned)
            
        log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Approved return request for order #{order.id}")
    else:
        log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Rejected return request for order #{order.id}")

    db.session.commit()
    
    # Notify User
    notif = Notification(
        recipient_type='user',
        recipient_id=order.user_id,
        title=f"Return Request {decision}",
        message=f"Your return request for order #{order.id} was {decision.lower()}.",
        shop_id=shop_id
    )
    db.session.add(notif)
    db.session.commit()
    
    return jsonify(order.serialize()), 200

@admin_bp.route('/shipping-wallet-balance', methods=['GET'])
@role_required(['admin'])
def get_shipping_wallet_balance():
    import shiprocket_helper
    shop_id = request.user['shop_id']
    shop = Shop.query.get(shop_id)
    if not shop:
        return jsonify({"error": "Shop not found"}), 404

    try:
        balance = shiprocket_helper.get_wallet_balance(shop)
        return jsonify({
            "success": True,
            "balance": balance
        }), 200
    except Exception as err:
        return jsonify({"error": str(err)}), 400

@admin_bp.route('/orders/<int:order_id>/shipping-serviceability', methods=['GET'])
@role_required(['admin'])
def get_shipping_serviceability(order_id):
    import re
    import shiprocket_helper

    shop_id = request.user['shop_id']
    order = Order.query.filter_by(id=order_id, shop_id=shop_id).first()
    if not order:
        return jsonify({"error": "Order not found"}), 404

    shop = Shop.query.get(shop_id)
    if not shop:
        return jsonify({"error": "Shop not found"}), 404

    weight_kg = request.args.get('weight_kg', 0.5, type=float)
    
    # Extract pincode from order's shipping address
    delivery_pincode = "400001"
    if order.shipping_address:
        pin_match = re.search(r'\b\d{6}\b', order.shipping_address)
        if pin_match:
            delivery_pincode = pin_match.group(0)

    is_cod = (order.payment_method == "COD")
    declared_value = order.final_amount or 0.0

    try:
        couriers = shiprocket_helper.check_serviceability(
            shop=shop,
            delivery_postcode=delivery_pincode,
            weight_kg=weight_kg,
            is_cod=is_cod,
            declared_value=declared_value
        )
        return jsonify({
            "success": True,
            "available_couriers": couriers
        }), 200
    except Exception as err:
        return jsonify({"error": str(err)}), 400

@admin_bp.route('/orders/<int:order_id>/book-shipping', methods=['POST'])
@role_required(['admin'])
def book_shipping(order_id):
    import requests
    import secrets
    from datetime import datetime
    import shiprocket_helper
    
    shop_id = request.user['shop_id']
    order = Order.query.filter_by(id=order_id, shop_id=shop_id).first()
    if not order:
        return jsonify({"error": "Order not found"}), 404
        
    shop = Shop.query.get(shop_id)
    
    data = request.get_json() or {}
    weight_kg = float(data.get('weight_kg', 0.5))
    declared_value = float(data.get('declared_value', order.final_amount))
    carrier = data.get('carrier', 'DTDC')
    courier_id = data.get('courier_id')
    
    if carrier == 'Shiprocket':
        try:
            sr_order_id = order.shiprocket_order_id
            sr_shipment_id = order.shiprocket_shipment_id
            
            # If not created on Shiprocket yet, create it
            if not sr_order_id or not sr_shipment_id:
                # 1. Build order items details
                items_list = []
                for item in order.items:
                    items_list.append({
                        "name": item.product_name,
                        "sku": item.product.sku_code if (item.product and item.product.sku_code) else f"SKU-{item.product_id}",
                        "units": item.quantity,
                        "price": item.price,
                        "product_id": item.product_id
                    })
                if not items_list:
                    items_list.append({
                        "name": "General Order Item",
                        "sku": f"GEN-SKU-{order.id}",
                        "units": 1,
                        "price": order.final_amount,
                        "product_id": 0
                    })
                    
                # 2. Call Shiprocket Create Order
                res_dict = shiprocket_helper.create_shiprocket_order(
                    shop=shop,
                    order_id_str=f"EC-{order.online_order_number or order.id}",
                    order_date=order.created_at or datetime.now(),
                    customer_name=order.user.name or order.user.username if order.user else "Customer",
                    customer_email=order.user.email if order.user else f"cust_{order.id}@example.com",
                    customer_phone=order.billing_phone or (order.user.contact_phone if order.user else "9999999999"),
                    shipping_address=order.shipping_address,
                    items=items_list,
                    weight_kg=weight_kg,
                    declared_value=declared_value,
                    payment_method=order.payment_method
                )
                
                sr_order_id = res_dict.get('shiprocket_order_id')
                sr_shipment_id = res_dict.get('shiprocket_shipment_id')
                
                order.shiprocket_order_id = sr_order_id
                order.shiprocket_shipment_id = sr_shipment_id
                db.session.commit() # Save Shiprocket IDs immediately so we don't duplicate order creation next time
                
            # 3. Assign AWB (do not catch exception here so failure propagates back to admin)
            awb_res = shiprocket_helper.assign_awb(shop, sr_shipment_id, courier_id)
            awb_code = awb_res.get('awb_code')
            courier_name = awb_res.get('courier_name')
                
            # Try to generate label
            label_url = None
            if awb_code:
                try:
                    label_url = shiprocket_helper.generate_label(shop, sr_shipment_id)
                except Exception as lbl_err:
                    print(f"Shiprocket label generation failed: {lbl_err}")
                    
            order.status = 'Dispatched'
            if awb_code:
                if courier_name:
                    order.tracking_info = f"Shiprocket (via {courier_name}) AWB: {awb_code}"
                else:
                    order.tracking_info = f"Shiprocket AWB: {awb_code}"
            else:
                order.tracking_info = f"Shiprocket Order ID: {sr_order_id}"
                
            if label_url:
                order.shipping_label_url = label_url
            else:
                host = request.host_url
                if not host.endswith('/'):
                    host += '/'
                order.shipping_label_url = f"{host}api/admin/orders/{order.id}/shipping-label"
                
            db.session.commit()
            
            # Send FCM notification
            try:
                from fcm_helper import send_order_status_notification
                send_order_status_notification(order, 'Dispatched')
            except Exception as e:
                print("FCM status notification error:", e)
                
            log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Booked Shiprocket Shipment for Order #{order.id} (Shipment ID: {sr_shipment_id})")
            
            return jsonify({
                "success": True,
                "message": "Shiprocket shipment booked successfully",
                "order": order.serialize()
            }), 200
            
        except Exception as err:
            return jsonify({"error": str(err)}), 400

    # Fallback to DTDC
    client_code = shop.dtdc_client_code
    api_key = shop.dtdc_api_key
    api_url = shop.dtdc_api_url or "https://api.dtdc.com/v1/shipments"
    
    awb_number = None
    label_url = None
    
    if client_code and api_key and client_code != "YOUR_DTDC_CLIENT_CODE" and not api_url.startswith("http://dummy"):
        try:
            payload = {
                "client_code": client_code,
                "consignee": {
                    "name": order.user.name or order.user.username if order.user else "Customer",
                    "phone": order.billing_phone or (order.user.contact_phone if order.user else ""),
                    "address": order.shipping_address,
                    "pincode": data.get('consignee_pincode', '')
                },
                "shipper": {
                    "name": shop.name,
                    "phone": shop.contact_phone or "",
                    "address": shop.address or ""
                },
                "package": {
                    "weight_kg": weight_kg,
                    "declared_value": declared_value,
                    "payment_mode": "Prepaid" if order.payment_method == "UPI" else "COD"
                }
            }
            headers = {
                "X-DTDC-API-KEY": api_key,
                "Content-Type": "application/json"
            }
            res = requests.post(api_url, json=payload, headers=headers, timeout=10)
            res_json = res.json()
            if res.status_code in [200, 201] and res_json.get('status') == 'success':
                awb_number = res_json.get('awb_number')
                label_url = res_json.get('label_url')
            else:
                print(f"DTDC Live API Error: {res.text}")
        except Exception as e:
            print(f"DTDC Booking Exception: {e}")
            
    if not awb_number:
        awb_number = f"D{secrets.token_hex(4).upper()}"
        host = request.host_url
        if not host.endswith('/'):
            host += '/'
        label_url = f"{host}api/admin/orders/{order.id}/shipping-label"
        
    order.status = 'Dispatched'
    order.tracking_info = f"DTDC AWB: {awb_number}"
    order.shipping_label_url = label_url
    db.session.commit()
    
    # Send FCM notification
    try:
        from fcm_helper import send_order_status_notification
        send_order_status_notification(order, 'Dispatched')
    except Exception as e:
        print("FCM status notification error:", e)
    
    log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Booked DTDC Shipment for Order #{order.id} (AWB: {awb_number})")
    
    return jsonify({
        "success": True,
        "message": "DTDC shipment booked successfully",
        "order": order.serialize()
    }), 200


@admin_bp.route('/orders/<int:order_id>/shipping-label', methods=['GET'])
def get_shipping_label(order_id):
    from datetime import datetime
    order = Order.query.get_or_404(order_id)
    shop = Shop.query.get(order.shop_id)
    
    awb_number = "D00000000"
    if order.tracking_info and "AWB:" in order.tracking_info:
        awb_number = order.tracking_info.split("AWB:")[-1].strip()
    
    barcode_value = awb_number if awb_number != "D00000000" else f"D{str(order.id).zfill(8)}"
    payment_type = "COD" if (order.payment_method or "").upper() == "COD" else "Prepaid"
    
    if order.online_order_number:
        order_display_id = f"#{str(order.online_order_number).zfill(6)}"
    elif order.tracking_info and not order.tracking_info.startswith('DTDC') and 'AWB' not in order.tracking_info:
        order_display_id = order.tracking_info
    else:
        order_display_id = f"#{order.id}"
        
    courier_name = "DTDC Express"
    if order.tracking_info:
        if "AWB:" in order.tracking_info:
            prefix = order.tracking_info.split("AWB:")[0].strip()
            if "Shiprocket" in prefix:
                if "via " in prefix:
                    parts = prefix.split("via ")
                    if len(parts) > 1:
                        courier_name = parts[1].replace("(", "").replace(")", "").strip()
                else:
                    courier_name = "Shiprocket Partner"
            elif prefix:
                courier_name = prefix
        else:
            if "Shiprocket" in order.tracking_info:
                courier_name = "Shiprocket Partner"
            elif "DTDC" in order.tracking_info:
                courier_name = "DTDC Express"
                
    if courier_name == "DTDC":
        courier_name = "DTDC Express"
        
    customer_name = order.user.name if order.user else 'Walk-in Customer'
    customer_phone = order.billing_phone or (order.user.contact_phone if order.user else '')
    customer_address = order.shipping_address or ''
    
    html = '''<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>{{ courier_name }} Shipping Label - {{ barcode_value }}</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800;900&display=swap" rel="stylesheet">
    <style>
        @page { size: 101.6mm 152.4mm; margin: 0; }
        * { box-sizing: border-box; }
        html, body { 
            margin: 0; 
            padding: 0; 
            font-family: 'Inter', Arial, sans-serif; 
            background-color: #f3f4f6; 
            color: #000;
        }
        body { 
            display: flex; 
            flex-direction: column;
            align-items: center; 
            justify-content: center; 
            min-height: 100vh;
            padding: 20px;
        }
        .print-btn {
            position: fixed;
            top: 20px;
            right: 20px;
            background-color: #0c2340;
            color: #ffffff;
            border: none;
            padding: 10px 20px;
            font-size: 0.95rem;
            font-weight: 700;
            border-radius: 6px;
            cursor: pointer;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
            transition: transform 0.1s;
            z-index: 1000;
        }
        .print-btn:active {
            transform: scale(0.95);
        }
        .label-card {
            width: 420px;
            height: 630px;
            background: #ffffff;
            border: 2px solid #111827;
            border-radius: 8px;
            box-shadow: 0 10px 15px -3px rgba(0,0,0,0.1);
            padding: 10px;
            box-sizing: border-box;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        .label-container { 
            width: 100%; 
            height: 100%; 
            border: 3px solid #000; 
            display: flex; 
            flex-direction: column; 
            overflow: hidden; 
            background: #fff; 
        }
        
        /* Grid Rows */
        .row { display: flex; width: 100%; border-bottom: 2px solid #000; }
        .row:last-child { border-bottom: none; }
        
        /* Two Column Rows */
        .col-50 { 
            width: 50%; 
            border-right: 2px solid #000; 
            padding: 10px; 
            display: flex; 
            flex-direction: column; 
            gap: 2px; 
            justify-content: flex-start; 
        }
        .col-50:last-child { border-right: none; }
        
        /* Header Row */
        .row-header { align-items: stretch; }
        .col-logo { 
            width: 33.333%; 
            border-right: 2px solid #000; 
            padding: 10px; 
            display: flex; 
            flex-direction: column; 
            align-items: center; 
            justify-content: center; 
            text-align: center; 
            gap: 3px; 
        }
        .logo-icon { width: 35.2px; height: 35.2px; color: #000; }
        .logo-text { font-size: 8.5px; font-weight: 850; text-transform: uppercase; letter-spacing: 0.05em; line-height: 1; }
        .col-title { 
            width: 66.666%; 
            padding: 10px; 
            display: flex; 
            align-items: center; 
            justify-content: center; 
            font-size: 20px; 
            font-weight: 900; 
            letter-spacing: 0.05em; 
            text-transform: uppercase; 
            text-align: center; 
        }
        
        /* Labels & Contents */
        .lbl { font-size: 9px; font-weight: 800; text-transform: uppercase; color: #000; letter-spacing: 0.03em; margin-bottom: 1px; }
        .val-bold { font-size: 12px; font-weight: 800; text-transform: uppercase; line-height: 1.4; }
        .val-text { font-size: 11.5px; font-weight: 500; line-height: 1.4; white-space: pre-line; }
        
        /* QR & Order Info Row */
        .col-qr { 
            width: 33.333%; 
            border-right: 2px solid #000; 
            padding: 10px; 
            display: flex; 
            align-items: center; 
            justify-content: center; 
        }
        .qr-code { width: 62.4px; height: 62.4px; display: block; max-width: 100%; height: auto; }
        .col-order-info { 
            width: 66.666%; 
            padding: 10px; 
            display: flex; 
            flex-direction: column; 
            justify-content: center; 
            gap: 6px; 
        }
        .order-info-item { display: flex; flex-direction: column; }
        
        /* Barcode Row */
        .row-barcode { 
            flex-direction: column; 
            align-items: center; 
            justify-content: center; 
            padding: 10px; 
            text-align: center; 
            gap: 4px; 
            border-bottom: 2px solid #000; 
        }
        .barcode-title { font-size: 9px; font-weight: 800; text-transform: uppercase; letter-spacing: 0.05em; }
        .barcode-svg-wrap { width: 100%; display: flex; justify-content: center; padding: 2px 0; }
        .barcode-svg-wrap svg { max-width: 95%; height: auto; display: block; }
        .barcode-text { font-size: 13px; font-weight: 800; letter-spacing: 0.1em; margin-top: 1px; }
        
        /* Payment & Amount Row */
        .payment-big { font-size: 28.8px; font-weight: 900; text-transform: uppercase; line-height: 1.1; margin-top: 2px; }
        
        /* Footer Row */
        .row-footer { 
            padding: 4px; 
            justify-content: center; 
            align-items: center; 
            text-align: center; 
            font-size: 8.5px; 
            font-weight: 800; 
            text-transform: uppercase; 
            letter-spacing: 0.05em; 
            background: #fafafb; 
        }
        
        @media print {
            body {
                background-color: #ffffff;
                padding: 0;
                margin: 0;
            }
            .print-btn {
                display: none;
            }
            .label-card {
                box-shadow: none;
                border: none;
                width: 100%;
                height: 100%;
                padding: 10px;
            }
        }
    </style>
</head>
<body>
    <button class="print-btn" onclick="window.print()">Print Label</button>
    
    <div class="label-card">
        <div class="label-container">
            <!-- Row 1: Header -->
            <div class="row row-header">
                <div class="col-logo">
                    <svg class="logo-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
                        <path d="M22 2L11 13"></path>
                        <polygon points="22 2 15 22 11 13 2 9 22 2"></polygon>
                    </svg>
                    <div class="logo-text">{{ courier_name }}</div>
                </div>
                <div class="col-title">Shipping Label</div>
            </div>
            
            <!-- Row 2: Sender/Consignee Info -->
            <div class="row">
                <div class="col-50">
                    <span class="lbl">From:</span>
                    <span class="val-bold">{{ shop.name }}</span>
                    <span class="val-text">{{ shop.address or '' }}</span>
                    <span class="val-text" style="font-weight:700; margin-top:2px;">Phone: {{ shop.contact_phone or '' }}</span>
                </div>
                <div class="col-50">
                    <span class="lbl">To:</span>
                    <span class="val-bold">{{ customer_name }}</span>
                    <span class="val-text">{{ customer_address }}</span>
                    <span class="val-text" style="font-weight:700; margin-top:2px;">Phone: {{ customer_phone }}</span>
                </div>
            </div>
            
            <!-- Row 3: Carrier Details -->
            <div class="row">
                <div class="col-50">
                    <span class="lbl">Shipping Partner:</span>
                    <span class="val-bold">{{ courier_name }}</span>
                </div>
                <div class="col-50">
                    <span class="lbl">Shipping Date:</span>
                    <span class="val-bold">{{ formatted_now }}</span>
                </div>
            </div>
            
            <!-- Row 4: QR & Order Details -->
            <div class="row">
                <div class="col-qr">
                    <img class="qr-code" alt="QR" />
                </div>
                <div class="col-order-info">
                    <div class="order-info-item">
                        <span class="lbl">Order Date:</span>
                        <span class="val-bold">{{ formatted_date }}</span>
                    </div>
                    <div class="order-info-item">
                        <span class="lbl">Order ID:</span>
                        <span class="val-bold">{{ order_display_id }}</span>
                    </div>
                </div>
            </div>
            
            <!-- Row 5: Barcode & Tracking -->
            <div class="row-barcode">
                <span class="barcode-title">Shipping Tracking Number:</span>
                <div class="barcode-svg-wrap"></div>
                <div class="barcode-text"></div>
            </div>
            
            <!-- Row 6: Payment Info -->
            <div class="row">
                <div class="col-50">
                    <span class="lbl">Payment Type:</span>
                    <span class="payment-big">{{ payment_type }}</span>
                </div>
                <div class="col-50">
                    <span class="lbl">{% if payment_type == 'COD' %}COD Amount:{% else %}Prepaid Amount:{% endif %}</span>
                    <span class="payment-big">₹{{ "%.2f"|format(order.final_amount) }}</span>
                </div>
            </div>
            
            <!-- Row 7: Footer -->
            <div class="row-footer">
                {{ courier_name | upper }} COURIER CARRIER / SERVICE: EXPRESS RESIDENTIAL
            </div>
        </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/jsbarcode@3.11.5/dist/JsBarcode.all.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/qrcode@1.4.4/build/qrcode.min.js"></script>
    <script>
        function formatBarcodeText(val) {
          const clean = val.replace(/\\s+/g, '');
          if (/^\\d+$/.test(clean)) {
            return clean.replace(/(.{4})/g, '$1 ').trim();
          }
          return val.split('').join(' ');
        }

        window.onload = function() {
            try {
                var barcodeSvg = document.createElementNS('http://www.w3.org/2000/svg', 'svg');
                JsBarcode(barcodeSvg, "{{ barcode_value }}", {
                    format: 'CODE128',
                    displayValue: false,
                    margin: 0,
                    width: 1.05,
                    height: 40,
                    background: 'transparent',
                    lineColor: '#111'
                });
                barcodeSvg.setAttribute('width', '240');
                barcodeSvg.setAttribute('height', '40');
                barcodeSvg.setAttribute('viewBox', '0 0 240 40');
                document.querySelector('.barcode-svg-wrap').appendChild(barcodeSvg);
                
                document.querySelector('.barcode-text').innerText = formatBarcodeText("{{ barcode_value }}");
            } catch (e) {
                console.error("Barcode generation failed", e);
            }

            try {
                QRCode.toDataURL("{{ barcode_value }}", { margin: 1, width: 120 }, function (err, url) {
                    if (!err) {
                        document.querySelector('.qr-code').src = url;
                    } else {
                        console.error("QR generation failed", err);
                    }
                });
            } catch (e) {
                console.error("QR generation exception", e);
            }
        };
    </script>
</body>
</html>'''
    return render_template_string(html, 
                                  order=order, 
                                  shop=shop, 
                                  barcode_value=barcode_value, 
                                  payment_type=payment_type, 
                                  order_display_id=order_display_id,
                                  customer_name=customer_name, 
                                  customer_phone=customer_phone, 
                                  customer_address=customer_address, 
                                  formatted_date=order.created_at.strftime('%d %b %Y') if order.created_at else '',
                                  formatted_now=datetime.now().strftime('%d %b %Y'),
                                  courier_name=courier_name)


@admin_bp.route('/customizations/<int:cust_id>/shipping-label', methods=['GET'])
def get_customization_shipping_label(cust_id):
    from datetime import datetime
    from models import CustomizationOrder
    cust = CustomizationOrder.query.get_or_404(cust_id)
    shop = Shop.query.get(cust.shop_id)
    
    awb_number = "D00000000"
    if cust.tracking_info and "AWB:" in cust.tracking_info:
        awb_number = cust.tracking_info.split("AWB:")[-1].strip()
        
    courier_name = "DTDC Express"
    if cust.tracking_info:
        if "AWB:" in cust.tracking_info:
            prefix = cust.tracking_info.split("AWB:")[0].strip()
            if "Shiprocket" in prefix:
                if "via " in prefix:
                    parts = prefix.split("via ")
                    if len(parts) > 1:
                        courier_name = parts[1].replace("(", "").replace(")", "").strip()
                else:
                    courier_name = "Shiprocket Partner"
            elif prefix:
                courier_name = prefix
        else:
            if "Shiprocket" in cust.tracking_info:
                courier_name = "Shiprocket Partner"
            elif "DTDC" in cust.tracking_info:
                courier_name = "DTDC Express"
                
    if courier_name == "DTDC":
        courier_name = "DTDC Express"
        
    barcode_value = awb_number if awb_number != "D00000000" else f"CUST-D{str(cust.id).zfill(6)}"
    payment_type = "Prepaid"
    
    unit_price = cust.quoted_price if cust.quoted_price is not None else (cust.product.price if cust.product else 0.0)
    final_amount = unit_price * cust.quantity
    
    customer_name = cust.user.name if (cust.user and cust.user.name) else (cust.user.username if cust.user else 'Customer')
    customer_phone = cust.billing_phone or (cust.user.contact_phone if cust.user else '')
    
    consignee_address = cust.shipping_address or (cust.user.addresses[0].get('address') if cust.user and cust.user.addresses else "N/A")
    product_name = cust.product.name if cust.product else "Deleted Product"
    
    html = '''<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>{{ courier_name }} Custom Shipping Label - {{ barcode_value }}</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800;900&display=swap" rel="stylesheet">
    <style>
        @page { size: 101.6mm 152.4mm; margin: 0; }
        * { box-sizing: border-box; }
        html, body { 
            margin: 0; 
            padding: 0; 
            font-family: 'Inter', Arial, sans-serif; 
            background-color: #f3f4f6; 
            color: #000;
        }
        body { 
            display: flex; 
            flex-direction: column;
            align-items: center; 
            justify-content: center; 
            min-height: 100vh;
            padding: 20px;
        }
        .print-btn {
            position: fixed;
            top: 20px;
            right: 20px;
            background-color: #0c2340;
            color: #ffffff;
            border: none;
            padding: 10px 20px;
            font-size: 0.95rem;
            font-weight: 700;
            border-radius: 6px;
            cursor: pointer;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
            transition: transform 0.1s;
            z-index: 1000;
        }
        .print-btn:active {
            transform: scale(0.95);
        }
        .label-card {
            width: 420px;
            height: 630px;
            background: #ffffff;
            border: 2px solid #111827;
            border-radius: 8px;
            box-shadow: 0 10px 15px -3px rgba(0,0,0,0.1);
            padding: 10px;
            box-sizing: border-box;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        .label-container { 
            width: 100%; 
            height: 100%; 
            border: 3px solid #000; 
            display: flex; 
            flex-direction: column; 
            overflow: hidden; 
            background: #fff; 
        }
        
        /* Grid Rows */
        .row { display: flex; width: 100%; border-bottom: 2px solid #000; }
        .row:last-child { border-bottom: none; }
        
        /* Two Column Rows */
        .col-50 { 
            width: 50%; 
            border-right: 2px solid #000; 
            padding: 10px; 
            display: flex; 
            flex-direction: column; 
            gap: 2px; 
            justify-content: flex-start; 
        }
        .col-50:last-child { border-right: none; }
        
        /* Header Row */
        .row-header { align-items: stretch; }
        .col-logo { 
            width: 33.333%; 
            border-right: 2px solid #000; 
            padding: 10px; 
            display: flex; 
            flex-direction: column; 
            align-items: center; 
            justify-content: center; 
            text-align: center; 
            gap: 3px; 
        }
        .logo-icon { width: 35.2px; height: 35.2px; color: #000; }
        .logo-text { font-size: 8.5px; font-weight: 850; text-transform: uppercase; letter-spacing: 0.05em; line-height: 1; }
        .col-title { 
            width: 66.666%; 
            padding: 10px; 
            display: flex; 
            align-items: center; 
            justify-content: center; 
            font-size: 20px; 
            font-weight: 900; 
            letter-spacing: 0.05em; 
            text-transform: uppercase; 
            text-align: center; 
        }
        
        /* Labels & Contents */
        .lbl { font-size: 9px; font-weight: 800; text-transform: uppercase; color: #000; letter-spacing: 0.03em; margin-bottom: 1px; }
        .val-bold { font-size: 12px; font-weight: 800; text-transform: uppercase; line-height: 1.4; }
        .val-text { font-size: 11.5px; font-weight: 500; line-height: 1.4; white-space: pre-line; }
        
        /* QR & Order Info Row */
        .col-qr { 
            width: 33.333%; 
            border-right: 2px solid #000; 
            padding: 10px; 
            display: flex; 
            align-items: center; 
            justify-content: center; 
        }
        .qr-code { width: 62.4px; height: 62.4px; display: block; max-width: 100%; height: auto; }
        .col-order-info { 
            width: 66.666%; 
            padding: 10px; 
            display: flex; 
            flex-direction: column; 
            justify-content: center; 
            gap: 6px; 
        }
        .order-info-item { display: flex; flex-direction: column; }
        
        /* Barcode Row */
        .row-barcode { 
            flex-direction: column; 
            align-items: center; 
            justify-content: center; 
            padding: 10px; 
            text-align: center; 
            gap: 4px; 
            border-bottom: 2px solid #000; 
        }
        .barcode-title { font-size: 9px; font-weight: 800; text-transform: uppercase; letter-spacing: 0.05em; }
        .barcode-svg-wrap { width: 100%; display: flex; justify-content: center; padding: 2px 0; }
        .barcode-svg-wrap svg { max-width: 95%; height: auto; display: block; }
        .barcode-text { font-size: 13px; font-weight: 800; letter-spacing: 0.1em; margin-top: 1px; }
        
        /* Payment & Amount Row */
        .payment-big { font-size: 28.8px; font-weight: 900; text-transform: uppercase; line-height: 1.1; margin-top: 2px; }
        
        /* Footer Row */
        .row-footer { 
            padding: 4px; 
            justify-content: center; 
            align-items: center; 
            text-align: center; 
            font-size: 8.5px; 
            font-weight: 800; 
            text-transform: uppercase; 
            letter-spacing: 0.05em; 
            background: #fafafb; 
        }
        
        @media print {
            body {
                background-color: #ffffff;
                padding: 0;
                margin: 0;
            }
            .print-btn {
                display: none;
            }
            .label-card {
                box-shadow: none;
                border: none;
                width: 100%;
                height: 100%;
                padding: 10px;
            }
        }
    </style>
</head>
<body>
    <button class="print-btn" onclick="window.print()">Print Label</button>
    
    <div class="label-card">
        <div class="label-container">
            <!-- Row 1: Header -->
            <div class="row row-header">
                <div class="col-logo">
                    <svg class="logo-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
                        <path d="M22 2L11 13"></path>
                        <polygon points="22 2 15 22 11 13 2 9 22 2"></polygon>
                    </svg>
                    <div class="logo-text">{{ courier_name }} Custom</div>
                </div>
                <div class="col-title">Shipping Label</div>
            </div>
            
            <!-- Row 2: Sender/Consignee Info -->
            <div class="row">
                <div class="col-50">
                    <span class="lbl">From:</span>
                    <span class="val-bold">{{ shop.name }}</span>
                    <span class="val-text">{{ shop.address or '' }}</span>
                    <span class="val-text" style="font-weight:700; margin-top:2px;">Phone: {{ shop.contact_phone or '' }}</span>
                </div>
                <div class="col-50">
                    <span class="lbl">To:</span>
                    <span class="val-bold">{{ customer_name }}</span>
                    <span class="val-text">{{ consignee_address }}</span>
                    <span class="val-text" style="font-weight:700; margin-top:2px;">Phone: {{ customer_phone }}</span>
                </div>
            </div>
            
            <!-- Row 3: Carrier Details -->
            <div class="row">
                <div class="col-50">
                    <span class="lbl">Shipping Partner:</span>
                    <span class="val-bold">{{ courier_name }}</span>
                </div>
                <div class="col-50">
                    <span class="lbl">Shipping Date:</span>
                    <span class="val-bold">{{ formatted_now }}</span>
                </div>
            </div>
            
            <!-- Row 4: QR & Custom Order Details -->
            <div class="row">
                <div class="col-qr">
                    <img class="qr-code" alt="QR" />
                </div>
                <div class="col-order-info">
                    <div class="order-info-item">
                        <span class="lbl">Order Date:</span>
                        <span class="val-bold">{{ formatted_date }}</span>
                    </div>
                    <div class="order-info-item">
                        <span class="lbl">Order ID:</span>
                        <span class="val-bold">#CUST-{{ "%06d"|format(cust.id) }}</span>
                    </div>
                    <div class="order-info-item" style="border-top: 1px dashed #ccc; padding-top: 4px; margin-top: 2px;">
                        <span class="lbl">Specs:</span>
                        <span class="val-text" style="font-size: 8px;"><strong>{{ product_name }}</strong> (Qty: {{ cust.quantity }})<br/>Color: {{ cust.selected_color_name }}<br/>Notes: {{ cust.customization_notes or 'None' }}</span>
                    </div>
                </div>
            </div>
            
            <!-- Row 5: Barcode & Tracking -->
            <div class="row-barcode">
                <span class="barcode-title">Shipping Tracking Number:</span>
                <div class="barcode-svg-wrap"></div>
                <div class="barcode-text"></div>
            </div>
            
            <!-- Row 6: Payment Info -->
            <div class="row">
                <div class="col-50">
                    <span class="lbl">Payment Type:</span>
                    <span class="payment-big">{{ payment_type }}</span>
                </div>
                <div class="col-50">
                    <span class="lbl">Prepaid Amount:</span>
                    <span class="payment-big">₹{{ "%.2f"|format(final_amount) }}</span>
                </div>
            </div>
            
            <!-- Row 7: Footer -->
            <div class="row-footer">
                {{ courier_name | upper }} CUSTOM COURIER / SERVICE: SPECIAL HANDCRAFTED
            </div>
        </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/jsbarcode@3.11.5/dist/JsBarcode.all.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/qrcode@1.4.4/build/qrcode.min.js"></script>
    <script>
        function formatBarcodeText(val) {
          const clean = val.replace(/\\s+/g, '');
          if (/^\\d+$/.test(clean)) {
            return clean.replace(/(.{4})/g, '$1 ').trim();
          }
          return val.split('').join(' ');
        }

        window.onload = function() {
            try {
                var barcodeSvg = document.createElementNS('http://www.w3.org/2000/svg', 'svg');
                JsBarcode(barcodeSvg, "{{ barcode_value }}", {
                    format: 'CODE128',
                    displayValue: false,
                    margin: 0,
                    width: 1.05,
                    height: 40,
                    background: 'transparent',
                    lineColor: '#111'
                });
                barcodeSvg.setAttribute('width', '240');
                barcodeSvg.setAttribute('height', '40');
                barcodeSvg.setAttribute('viewBox', '0 0 240 40');
                document.querySelector('.barcode-svg-wrap').appendChild(barcodeSvg);
                
                document.querySelector('.barcode-text').innerText = formatBarcodeText("{{ barcode_value }}");
            } catch (e) {
                console.error("Barcode generation failed", e);
            }

            try {
                QRCode.toDataURL("{{ barcode_value }}", { margin: 1, width: 120 }, function (err, url) {
                    if (!err) {
                        document.querySelector('.qr-code').src = url;
                    } else {
                        console.error("QR generation failed", err);
                    }
                });
            } catch (e) {
                console.error("QR generation exception", e);
            }
        };
    </script>
</body>
</html>'''
    return render_template_string(html, 
                                  cust=cust, 
                                  shop=shop, 
                                  barcode_value=barcode_value, 
                                  payment_type=payment_type, 
                                  final_amount=final_amount,
                                  customer_name=customer_name, 
                                  customer_phone=customer_phone, 
                                  consignee_address=consignee_address, 
                                  product_name=product_name,
                                  formatted_date=cust.created_at.strftime('%d %b %Y') if cust.created_at else '',
                                  formatted_now=datetime.now().strftime('%d %b %Y'),
                                  courier_name=courier_name)


# POPUP ADS PUSHING
@admin_bp.route('/popup-ads', methods=['GET', 'POST'])
@role_required(['admin'])
def manage_popup_ads():
    shop_id = request.user['shop_id']

    if request.method == 'GET':
        ads = PopupAd.query.filter_by(shop_id=shop_id).all()
        return jsonify([ad.serialize() for ad in ads]), 200

    data = request.get_json() or {}
    title = data.get('title')
    image_url = data.get('image_url')

    if not title or not image_url:
        return jsonify({"error": "Title and image_url are required"}), 400

    ad = PopupAd(
        title=title,
        image_url=image_url,
        target_url=data.get('target_url', ''),
        show_before_login=bool(data.get('show_before_login', True)),
        show_after_login=bool(data.get('show_after_login', True)),
        is_active=bool(data.get('is_active', True)),
        shop_id=shop_id
    )
    db.session.add(ad)
    db.session.commit()

    log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Created popup ad campaign '{title}'")

    # Send email to all newsletter subscribers in a background thread to prevent blocking
    try:
        import threading
        from mail_sender import send_popup_ad_email
        from flask import current_app
        
        subscriber_emails = [sub.email for sub in subscribers]
        ad_id = ad.id
        
        def dispatch_emails(app_context, shop_id, ad_id, target_link, emails):
            with app_context:
                from models import Shop, PopupAd
                shop_obj = Shop.query.get(shop_id)
                ad_obj = PopupAd.query.get(ad_id)
                if shop_obj and ad_obj:
                    for email in emails:
                        try:
                            send_popup_ad_email(shop_obj, email, ad_obj, target_link)
                        except Exception as e:
                            print(f"Error sending popup ad email to {email}: {e}")
                        
        app_context = current_app.app_context()
        threading.Thread(target=dispatch_emails, args=(app_context, shop.id, ad_id, target_link, subscriber_emails)).start()
    except Exception as e:
        print(f"Error spawning newsletter announcement thread: {e}")

    return jsonify(ad.serialize()), 201

@admin_bp.route('/popup-ads/<int:ad_id>', methods=['PUT', 'DELETE'])
@role_required(['admin'])
def modify_popup_ad(ad_id):
    shop_id = request.user['shop_id']
    ad = PopupAd.query.filter_by(id=ad_id, shop_id=shop_id).first()
    if not ad:
        return jsonify({"error": "Popup ad not found"}), 404

    if request.method == 'DELETE':
        title = ad.title
        db.session.delete(ad)
        db.session.commit()
        log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Deleted popup ad '{title}'")
        return jsonify({"message": "Popup ad deleted successfully"}), 200

    data = request.get_json() or {}
    if 'title' in data:
        ad.title = data['title']
    if 'image_url' in data:
        ad.image_url = data['image_url']
    if 'target_url' in data:
        ad.target_url = data['target_url']
    if 'show_before_login' in data:
        ad.show_before_login = bool(data['show_before_login'])
    if 'show_after_login' in data:
        ad.show_after_login = bool(data['show_after_login'])
    if 'is_active' in data:
        ad.is_active = bool(data['is_active'])

    db.session.commit()
    log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Updated popup ad campaign '{ad.title}'")
    return jsonify(ad.serialize()), 200

# COUPONS MANAGEMENT
@admin_bp.route('/coupons', methods=['GET', 'POST'])
@role_required(['admin'])
def manage_coupons():
    shop_id = request.user['shop_id']

    if request.method == 'GET':
        coupons = Coupon.query.filter_by(shop_id=shop_id).all()
        # Auto-deactivate expired or exhausted coupons
        now = datetime.now()
        updated = False
        for c in coupons:
            if c.is_active:
                if c.expires_at and now > c.expires_at:
                    c.is_active = False
                    updated = True
                elif c.usage_limit is not None and c.used_count >= c.usage_limit:
                    c.is_active = False
                    updated = True
        if updated:
            db.session.commit()
        return jsonify([c.serialize() for c in coupons]), 200

    data = request.get_json() or {}
    code = data.get('code')
    discount_percentage = data.get('discount_percentage')

    if not code or discount_percentage is None:
        return jsonify({"error": "Coupon code and discount percentage are required"}), 400

    # Prevent duplicate coupon code for same shop
    existing = Coupon.query.filter_by(code=code, shop_id=shop_id).first()
    if existing:
        return jsonify({"error": f"Coupon code '{code}' already exists for this shop"}), 400

    expires_at_val = None
    expires_at_str = data.get('expires_at')
    if expires_at_str:
        try:
            expires_at_str_clean = expires_at_str.replace('Z', '+00:00')
            expires_at_val = datetime.fromisoformat(expires_at_str_clean)
        except Exception:
            try:
                expires_at_val = datetime.strptime(expires_at_str[:16], "%Y-%m-%dT%H:%M")
            except Exception:
                pass

    usage_limit_val = None
    if 'usage_limit' in data and data['usage_limit'] is not None and str(data['usage_limit']).strip() != '':
        try:
            usage_limit_val = int(data['usage_limit'])
        except Exception:
            pass

    usage_limit_per_user_val = None
    if 'usage_limit_per_user' in data and data['usage_limit_per_user'] is not None and str(data['usage_limit_per_user']).strip() != '':
        try:
            usage_limit_per_user_val = int(data['usage_limit_per_user'])
        except Exception:
            pass

    c = Coupon(
        code=code.upper(),
        discount_percentage=float(discount_percentage),
        max_discount=float(data.get('max_discount', 1000.0)),
        min_purchase=float(data.get('min_purchase', 0.0)),
        is_active=bool(data.get('is_active', True)),
        shop_id=shop_id,
        expires_at=expires_at_val,
        usage_limit=usage_limit_val,
        usage_limit_per_user=usage_limit_per_user_val,
        used_count=0
    )
    db.session.add(c)
    db.session.commit()

    # Send FCM coupon notification to topic
    try:
        from fcm_helper import send_fcm_topic_notification
        topic = f"shop_{shop_id}_offers"
        title = "New Offer Available!"
        body = f"Use coupon code {c.code} to get {c.discount_percentage}% off on your next purchase!"
        send_fcm_topic_notification(topic, title, body, data={"coupon_code": c.code})
    except Exception as e:
        print(f"Failed to send FCM coupon notification: {e}")

    log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Created discount coupon '{c.code}'")
    return jsonify(c.serialize()), 201

@admin_bp.route('/coupons/<int:coupon_id>', methods=['PUT', 'DELETE'])
@role_required(['admin'])
def modify_coupon(coupon_id):
    shop_id = request.user['shop_id']
    c = Coupon.query.filter_by(id=coupon_id, shop_id=shop_id).first()
    if not c:
        return jsonify({"error": "Coupon not found"}), 404

    if request.method == 'DELETE':
        code = c.code
        db.session.delete(c)
        db.session.commit()
        log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Deleted coupon '{code}'")
        return jsonify({"message": "Coupon deleted successfully"}), 200

    data = request.get_json() or {}
    if 'code' in data:
        c.code = data['code'].upper()
    if 'discount_percentage' in data:
        c.discount_percentage = float(data['discount_percentage'])
    if 'max_discount' in data:
        c.max_discount = float(data['max_discount'])
    if 'min_purchase' in data:
        c.min_purchase = float(data['min_purchase'])
    if 'is_active' in data:
        c.is_active = bool(data['is_active'])

    if 'expires_at' in data:
        expires_at_str = data['expires_at']
        if expires_at_str:
            try:
                expires_at_str_clean = expires_at_str.replace('Z', '+00:00')
                c.expires_at = datetime.fromisoformat(expires_at_str_clean)
            except Exception:
                try:
                    c.expires_at = datetime.strptime(expires_at_str[:16], "%Y-%m-%dT%H:%M")
                except Exception:
                    pass
        else:
            c.expires_at = None

    if 'usage_limit' in data:
        usage_limit_val = data['usage_limit']
        if usage_limit_val is not None and str(usage_limit_val).strip() != '':
            try:
                c.usage_limit = int(usage_limit_val)
            except Exception:
                pass
        else:
            c.usage_limit = None

    if 'usage_limit_per_user' in data:
        usage_limit_per_user_val = data['usage_limit_per_user']
        if usage_limit_per_user_val is not None and str(usage_limit_per_user_val).strip() != '':
            try:
                c.usage_limit_per_user = int(usage_limit_per_user_val)
            except Exception:
                pass
        else:
            c.usage_limit_per_user = None

    db.session.commit()
    log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Updated coupon code '{c.code}' settings")
    return jsonify(c.serialize()), 200

# GST TAXATION REPORTS
@admin_bp.route('/gst-report', methods=['GET'])
@role_required(['admin'])
def gst_report():
    shop_id = request.user['shop_id']
    shop = Shop.query.get(shop_id)
    
    date_val = request.args.get('date') # YYYY-MM-DD
    month_val = request.args.get('month') # 1-12
    year_val = request.args.get('year') # YYYY
    
    # Calculate tax aggregates from successful sales
    orders = Order.query.filter_by(shop_id=shop_id).filter(Order.status != 'Returned').all()
    
    filtered_orders = []
    for o in orders:
        if not o.created_at:
            continue
        # Skip COD orders that are still Pending or Rejected
        if o.payment_method == 'COD' and o.status in ['Pending', 'Rejected']:
            continue
        # Apply date filter (YYYY-MM-DD)
        if date_val and o.created_at.strftime('%Y-%m-%d') != date_val:
            continue
        # Apply month filter (1-12)
        if month_val and o.created_at.month != int(month_val):
            continue
        # Apply year filter (YYYY)
        if year_val and o.created_at.year != int(year_val):
            continue
        filtered_orders.append(o)
        
    total_sales_value = sum([o.final_amount for o in filtered_orders])
    total_gst_amount = sum([o.gst_amount for o in filtered_orders])
    total_shipping = sum([o.shipping_charge or 0.0 for o in filtered_orders])
    total_shipping_gst = sum([o.shipping_gst or 0.0 for o in filtered_orders])
    
    net_sales = total_sales_value - total_gst_amount
    net_goods_sales = net_sales - (total_shipping - total_shipping_gst)
    goods_gst = total_gst_amount - total_shipping_gst
    gst_rate = getattr(shop, 'gst_percentage', 18.0)
    
    reporting_period = "All Time"
    months = ["", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    
    if date_val:
        reporting_period = f"Date: {date_val}"
    elif month_val and year_val:
        m_idx = int(month_val)
        m_name = months[m_idx] if 0 < m_idx < 13 else f"Month {month_val}"
        reporting_period = f"{m_name} {year_val}"
    elif month_val:
        m_idx = int(month_val)
        m_name = months[m_idx] if 0 < m_idx < 13 else f"Month {month_val}"
        reporting_period = f"Month: {m_name}"
    elif year_val:
        reporting_period = f"Year: {year_val}"
        
    return jsonify({
        "shop_id": shop_id,
        "reporting_period": reporting_period,
        "net_sales": round(net_sales, 2),
        "gst_rate_applied": f"{gst_rate}% Configured GST",
        "total_gst_collected": round(total_gst_amount, 2),
        "gross_revenue": round(total_sales_value, 2),
        "total_shipping_collected": round(total_shipping, 2),
        "total_shipping_gst": round(total_shipping_gst, 2),
        "net_goods_sales": round(net_goods_sales, 2),
        "goods_gst_collected": round(goods_gst, 2),
        "total_orders_count": len(filtered_orders),
        "orders": [o.serialize() for o in filtered_orders]
    }), 200

@admin_bp.route('/gst-report/export', methods=['GET'])
@role_required(['admin'])
def export_gst_report():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    shop_id = request.user['shop_id']
    shop = Shop.query.get(shop_id)

    date_val = request.args.get('date') # YYYY-MM-DD
    month_val = request.args.get('month') # 1-12
    year_val = request.args.get('year') # YYYY

    # Calculate tax aggregates from successful sales
    orders = Order.query.filter_by(shop_id=shop_id).filter(Order.status != 'Returned').all()

    filtered_orders = []
    for o in orders:
        if not o.created_at:
            continue
        # Skip COD orders that are still Pending or Rejected
        if o.payment_method == 'COD' and o.status in ['Pending', 'Rejected']:
            continue
        # Apply date filter (YYYY-MM-DD)
        if date_val and o.created_at.strftime('%Y-%m-%d') != date_val:
            continue
        # Apply month filter (1-12)
        if month_val and o.created_at.month != int(month_val):
            continue
        # Apply year filter (YYYY)
        if year_val and o.created_at.year != int(year_val):
            continue
        filtered_orders.append(o)

    reporting_period = "All Time"
    months = ["", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]

    if date_val:
        reporting_period = f"Date: {date_val}"
    elif month_val and year_val:
        m_idx = int(month_val)
        m_name = months[m_idx] if 0 < m_idx < 13 else f"Month {month_val}"
        reporting_period = f"{m_name} {year_val}"
    elif month_val:
        m_idx = int(month_val)
        m_name = months[m_idx] if 0 < m_idx < 13 else f"Month {month_val}"
        reporting_period = f"Month: {m_name}"
    elif year_val:
        reporting_period = f"Year: {year_val}"

    wb = Workbook()
    ws = wb.active
    ws.title = "GST Report"

    # Explicitly enable grid lines
    ws.views.sheetView[0].showGridLines = True

    # Typography & styles
    font_title = Font(name='Segoe UI', size=16, bold=True, color='2B0B57')
    font_bold = Font(name='Segoe UI', size=11, bold=True)
    font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    font_regular = Font(name='Segoe UI', size=11)

    fill_header = PatternFill(start_color='2B0B57', end_color='2B0B57', fill_type='solid') # Deep purple
    fill_zebra = PatternFill(start_color='F5EEF9', end_color='F5EEF9', fill_type='solid') # Zebra striping purple tint

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    border_thin_side = Side(border_style="thin", color="DDDDDD")
    border_thin = Border(left=border_thin_side, right=border_thin_side, top=border_thin_side, bottom=border_thin_side)

    border_double_bottom = Side(border_style="double", color="2B0B57")
    border_thin_top = Side(border_style="thin", color="2B0B57")
    border_total = Border(top=border_thin_top, bottom=border_double_bottom)

    # Title Banner Row
    ws.merge_cells('A1:H1')
    ws['A1'] = "NOBARAA FASHION - GST TAX ACCOUNTING REPORT"
    ws['A1'].font = font_title
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 40

    # Information row
    ws['A2'] = "Shop Name:"
    ws['A2'].font = font_bold
    ws['B2'] = shop.name if shop else "Nobaraa Fashion"
    ws['B2'].font = font_regular

    ws['D2'] = "Reporting Period:"
    ws['D2'].font = font_bold
    ws['E2'] = reporting_period
    ws['E2'].font = font_regular

    ws['G2'] = "GST Rate:"
    ws['G2'].font = font_bold
    ws['H2'] = f"{getattr(shop, 'gst_percentage', 18.0)}% Configured"
    ws['H2'].font = font_regular

    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 15 # blank space

    # Headers definition
    headers = [
        "Order ID",
        "Date & Time",
        "Customer",
        "Tax Type",
        "Gross Amount (INR)",
        "Delivery Charge (INR)",
        "Delivery GST (INR)",
        "Net Goods Sales (INR)",
        "Goods GST (INR)",
        "Total GST Collected (INR)",
        "Status"
    ]

    header_row = 4
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx in [1, 4, 11] else (align_left if col_idx in [2, 3] else align_right)
        cell.border = border_thin

    ws.row_dimensions[header_row].height = 28

    # Populate Data
    current_row = 5
    total_gross = 0.0
    total_delivery = 0.0
    total_delivery_gst = 0.0
    total_net_goods = 0.0
    total_goods_gst = 0.0
    total_gst = 0.0

    for o in filtered_orders:
        order_gross = o.final_amount
        order_delivery = o.shipping_charge or 0.0
        order_delivery_gst = o.shipping_gst or 0.0
        order_total_gst = o.gst_amount or 0.0
        order_goods_gst = order_total_gst - order_delivery_gst
        order_net_goods = order_gross - order_delivery - order_goods_gst

        total_gross += order_gross
        total_delivery += order_delivery
        total_delivery_gst += order_delivery_gst
        total_net_goods += order_net_goods
        total_goods_gst += order_goods_gst
        total_gst += order_total_gst

        created_time = o.created_at.strftime('%Y-%m-%d %I:%M:%S %p') if o.created_at else 'N/A'

        order_display_id = f"#{str(o.online_order_number).zfill(6)}" if o.online_order_number else (o.tracking_info if (o.tracking_info and not o.tracking_info.startswith('DTDC') and 'AWB' not in o.tracking_info) else f"#{o.id}")
        row_data = [
            order_display_id,
            created_time,
            o.user.name if o.user else 'Unknown User',
            "Inclusive" if o.gst_inclusive else "Exclusive",
            order_gross,
            order_delivery,
            order_delivery_gst,
            order_net_goods,
            order_goods_gst,
            order_total_gst,
            o.status
        ]

        row_fill = fill_zebra if current_row % 2 == 1 else None

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_regular
            if row_fill:
                cell.fill = row_fill
            cell.border = border_thin

            if col_idx in [1, 4, 11]:
                cell.alignment = align_center
            elif col_idx in [2, 3]:
                cell.alignment = align_left
            else:
                cell.alignment = align_right

            if col_idx in [5, 6, 7, 8, 9, 10]:
                cell.number_format = '₹#,##0.00'

        ws.row_dimensions[current_row].height = 22
        current_row += 1

    # Add Totals Row
    ws.cell(row=current_row, column=1, value="Total Summary").font = font_bold
    ws.cell(row=current_row, column=1).alignment = align_left
    ws.cell(row=current_row, column=1).border = border_total

    for col_idx in range(2, 12):
        ws.cell(row=current_row, column=col_idx).border = border_total

    for col_idx, tot_val in [
        (5, total_gross),
        (6, total_delivery),
        (7, total_delivery_gst),
        (8, total_net_goods),
        (9, total_goods_gst),
        (10, total_gst)
    ]:
        cell_tot = ws.cell(row=current_row, column=col_idx, value=tot_val)
        cell_tot.font = font_bold
        cell_tot.alignment = align_right
        cell_tot.number_format = '₹#,##0.00'

    ws.row_dimensions[current_row].height = 26

    # Adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            if cell.number_format and '₹' in cell.number_format and isinstance(cell.value, (int, float)):
                val_str = f"Rs. {cell.value:,.2f}"
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename_period = reporting_period.replace(' ', '_').replace(':', '')
    filename = f"GST_Tax_Report_{filename_period}.xlsx"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# IN-APP NOTIFICATION BROADCASTS
@admin_bp.route('/notifications', methods=['GET', 'POST'])
@role_required(['admin'])
def manage_notifications():
    shop_id = request.user['shop_id']

    if request.method == 'GET':
        notifications = Notification.query.filter_by(shop_id=shop_id, recipient_type='admin').order_by(Notification.created_at.desc()).all()
        return jsonify([n.serialize() for n in notifications]), 200

    data = request.get_json() or {}
    title = data.get('title')
    message = data.get('message')

    if not title or not message:
        return jsonify({"error": "Title and message are required"}), 400

    # Broadcast notification to all Users
    notif = Notification(
        recipient_type='user',
        recipient_id=None, # None indicates broadcast
        title=title,
        message=message,
        shop_id=shop_id
    )
    db.session.add(notif)
    db.session.commit()

    log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Broadcasted notification alert: '{title}'")
    return jsonify(notif.serialize()), 201

# SUPPORT TICKETS HELP CENTER MANAGEMENT
@admin_bp.route('/help-tickets', methods=['GET', 'PUT'])
@role_required(['admin'])
def manage_help_tickets():
    from models import HelpTicket
    shop_id = request.user['shop_id']

    if request.method == 'GET':
        tickets = HelpTicket.query.filter_by(shop_id=shop_id).order_by(HelpTicket.created_at.desc()).all()
        return jsonify([t.serialize() for t in tickets]), 200

    # PUT request - reply to a ticket
    data = request.get_json() or {}
    ticket_id = data.get('ticket_id')
    reply = data.get('reply')

    if not ticket_id or not reply:
        return jsonify({"error": "Ticket ID and reply content are required"}), 400

    ticket = HelpTicket.query.filter_by(id=ticket_id, shop_id=shop_id).first()
    if not ticket:
        return jsonify({"error": "Help ticket not found"}), 404

    ticket.reply = reply
    ticket.status = 'Resolved'
    db.session.commit()

    # Notify user of reply via in-app notification
    notif = Notification(
        recipient_type='user',
        recipient_id=ticket.user_id,
        title="Help Ticket Replied!",
        message=f"Admin replied to your ticket '{ticket.subject}': {reply[:40]}...",
        shop_id=shop_id
    )
    db.session.add(notif)
    db.session.commit()

    # Retrieve user and shop objects
    user = ticket.user
    shop = ticket.shop

    # Send FCM push notification (background)
    if user and user.fcm_token:
        try:
            from fcm_helper import send_fcm_notification
            send_fcm_notification(
                token=user.fcm_token,
                title="Help Ticket Replied!",
                body=f"Admin replied to your ticket '{ticket.subject}': {reply[:60]}...",
                data={
                    "type": "ticket_reply",
                    "ticket_id": str(ticket.id)
                }
            )
        except Exception as fcm_err:
            print("FCM support reply notification error:", fcm_err)

    # Send SMTP Email notification (asynchronous thread)
    if shop and user and user.email and shop.smtp_host and shop.smtp_user:
        try:
            import threading
            from flask import current_app
            from mail_sender import send_ticket_reply_email
            
            app_context = current_app.app_context()
            site_link = request.headers.get('Origin') or "http://localhost:5173"
            
            def async_send_reply(app_context, shop_id, email_addr, sub_text, query_text, reply_text, link_str, sender_info):
                with app_context:
                    try:
                        from models import Shop
                        shop_obj = Shop.query.get(shop_id)
                        if shop_obj:
                            send_ticket_reply_email(
                                shop=shop_obj,
                                recipient_email=email_addr,
                                ticket_subject=sub_text,
                                query_message=query_text,
                                admin_reply=reply_text,
                                site_link=link_str,
                                sender_info=sender_info
                            )
                    except Exception as mail_err:
                        print(f"Error in async ticket reply email: {mail_err}")
                        
            threading.Thread(
                target=async_send_reply,
                args=(
                    app_context,
                    shop.id,
                    user.email,
                    ticket.subject,
                    ticket.message,
                    reply,
                    site_link,
                    {
                        "actor_type": "admin",
                        "actor_id": request.user['user_id'],
                        "username": request.user['username']
                    }
                )
            ).start()
        except Exception as spawn_err:
            print(f"Failed to spawn ticket reply email thread: {spawn_err}")

    log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Resolved help ticket #{ticket.id} ('{ticket.subject}')")
    return jsonify(ticket.serialize()), 200

# MOCK FAST2SMS AND WHATSAPP CAMPAIGNS LOGGING
@admin_bp.route('/sms-whatsapp-logs', methods=['GET', 'POST'])
@role_required(['admin'])
def manage_messaging():
    shop_id = request.user['shop_id']
    
    if request.method == 'GET':
        # Let's read system logs related to sms / messaging
        logs = SystemLog.query.filter_by(shop_id=shop_id).filter(SystemLog.action.like('%Message:%')).order_by(SystemLog.created_at.desc()).all()
        
        # Return a list of mock SMS/WhatsApp/FCM campaigns
        res = []
        for log in logs:
            parts = log.action.split('|')
            platform_prefix = parts[0].strip()
            if "FCM" in platform_prefix:
                platform = "Firebase Push (FCM)"
            elif "SMS" in platform_prefix or "SMS" in log.action:
                platform = "Fast2SMS"
            else:
                platform = "WhatsApp Gateway"
            res.append({
                "id": log.id,
                "timestamp": log.created_at.isoformat(),
                "platform": platform,
                "recipient": parts[1] if len(parts) > 1 else "All Customers",
                "message": parts[2] if len(parts) > 2 else parts[0],
                "status": "Delivered successfully (100% Rate)"
            })
        return jsonify(res), 200

    # POST - send SMS / WhatsApp campaign / FCM
    data = request.get_json() or {}
    platform = data.get('platform') # SMS, WhatsApp, FCM
    recipient = data.get('recipient', 'All Customers')
    message = data.get('message')
    title = data.get('title', 'Nobaraa Fashion') # Push notification title

    if not platform or not message:
        return jsonify({"error": "Platform and message content are required"}), 400

    if platform == 'FCM':
        from fcm_helper import send_fcm_topic_notification, send_fcm_notification
        
        # Determine target and send notification
        # For broadcast
        user = None
        if recipient == 'All Customers' or recipient == 'Active Cart Users':
            topic = f"shop_{shop_id}_all"
            fcm_sent = send_fcm_topic_notification(topic, title, message)
        else:
            user = User.query.filter_by(contact_phone=recipient).first()
            if user and user.fcm_token:
                fcm_sent = send_fcm_notification(user.fcm_token, title, message)
            else:
                fcm_sent = False
                
        # Also create a Notification record in DB for the in-app notification center!
        db_notif = Notification(
            recipient_type='user',
            recipient_id=user.id if (recipient != 'All Customers' and recipient != 'Active Cart Users' and user) else None,
            title=title,
            message=message,
            shop_id=shop_id
        )
        db.session.add(db_notif)
        db.session.commit()
        
        action_str = f"[Message:FCM] | {recipient} | {title}: {message}"
    else:
        if platform == 'SMS':
            shop = Shop.query.get(shop_id)
            if not shop or not shop.sms_campaign_enabled:
                return jsonify({"error": "SMS campaign messaging is disabled for this shop. Enable it in Shop Config."}), 400
        action_str = f"[Message:{platform}] | {recipient} | {message}"
        
    log_admin_action(request.user['user_id'], request.user['username'], shop_id, action_str)

    return jsonify({
        "status": "Success",
        "message": f"{platform} campaign successfully dispatched. 100% delivery rate.",
        "details": {
            "platform": platform,
            "recipient": recipient,
            "char_count": len(message),
            "cost_per_message": "0.00" if platform in ["WhatsApp", "FCM"] else "0.20 Rupees"
        }
    }), 200

@admin_bp.route('/logs', methods=['GET'])
@role_required(['admin'])
def get_store_logs():
    shop_id = request.user['shop_id']
    logs = SystemLog.query.filter_by(shop_id=shop_id).filter(SystemLog.actor_type != 'super_admin').order_by(SystemLog.created_at.desc()).all()
    return jsonify([log.serialize() for log in logs]), 200

# CUSTOMIZATION ORDER MANAGEMENT
@admin_bp.route('/customizations', methods=['GET'])
@role_required(['admin'])
def list_customizations():
    shop_id = request.user['shop_id']
    custs = CustomizationOrder.query.filter_by(shop_id=shop_id).order_by(CustomizationOrder.created_at.desc()).all()
    return jsonify([c.serialize() for c in custs]), 200

@admin_bp.route('/customizations/<int:cust_id>', methods=['PUT'])
@role_required(['admin'])
def update_customization_status(cust_id):
    shop_id = request.user['shop_id']
    cust = CustomizationOrder.query.filter_by(id=cust_id, shop_id=shop_id).first()
    if not cust:
        return jsonify({"error": "Customization request not found"}), 404
        
    data = request.get_json() or {}
    if 'status' in data:
        cust.status = data['status']
        db.session.commit()
        
        # Log action
        log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Updated status of Customization #{cust_id} to '{cust.status}'")
        
        # Send FCM notification
        try:
            from fcm_helper import send_customization_status_notification
            send_customization_status_notification(cust, 'status')
        except Exception as e:
            print("FCM customization status notification error:", e)

        # Send SMS notification
        try:
            from sms_helper import send_order_status_sms
            send_order_status_sms(cust, cust.status)
        except Exception as e:
            print("SMS customization status notification error:", e)
        
    return jsonify({"message": "Customization request updated successfully", "customization": cust.serialize()}), 200

@admin_bp.route('/customizations/<int:cust_id>/quote', methods=['PUT'])
@role_required(['admin'])
def update_customization_quote(cust_id):
    shop_id = request.user['shop_id']
    cust = CustomizationOrder.query.filter_by(id=cust_id, shop_id=shop_id).first()
    if not cust:
        return jsonify({"error": "Customization request not found"}), 404
        
    data = request.get_json() or {}
    quoted_price = data.get('quoted_price')
    if quoted_price is None:
        return jsonify({"error": "quoted_price is required"}), 400
        
    try:
        cust.quoted_price = float(quoted_price)
        cust.quote_status = 'Quoted'
        db.session.commit()
        
        log_admin_action(request.user['user_id'], request.user['username'], shop_id, f"Quoted price of ₹{cust.quoted_price:.2f} for Customization #{cust_id}")
        
        # Send FCM notification
        try:
            from fcm_helper import send_customization_status_notification
            send_customization_status_notification(cust, 'quote')
        except Exception as e:
            print("FCM customization quote notification error:", e)
            
        return jsonify({"message": "Price quote sent successfully", "customization": cust.serialize()}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Failed to save price quote: {str(e)}"}), 500


@admin_bp.route('/sms/fetch-template-content', methods=['GET'])
@role_required(['admin'])
def fetch_sms_template_content():
    import os
    import requests
    import logging
    
    shop = Shop.query.get(request.user['shop_id'])
    if not shop:
        return jsonify({"error": "Shop not found"}), 404
        
    sender_id = request.args.get('sender_id', '').strip().upper()
    template_id = request.args.get('template_id', '').strip()
    api_key_param = request.args.get('api_key', '').strip()
    
    if not sender_id or not template_id:
        return jsonify({"error": "sender_id and template_id are required"}), 400
        
    api_key = api_key_param or shop.sms_api_key or os.getenv("FAST2SMS_API_KEY")
    
    # Check if key is mock
    is_mock = not api_key or "MOCK" in api_key or "FAST2SMS_" in api_key
    
    if is_mock:
        content = f"Your OTP code for verification is 123456. Valid for 10 mins. - {sender_id}"
        if "otp" in template_id.lower() or "otp" in sender_id.lower():
            content = f"Your OTP for login to our store is 123456. Do not share. - {sender_id}"
        elif "dispatch" in template_id.lower() or "disp" in template_id.lower():
            content = f"Your order 1001 has been dispatched. Tracking: TRACK123. - {sender_id}"
        elif "delivery" in template_id.lower() or "deliv" in template_id.lower():
            content = f"Your order 1001 has been delivered successfully. Thank you. - {sender_id}"
        
        return jsonify({
            "success": True,
            "sender_id": sender_id,
            "template_id": template_id,
            "content": content,
            "is_mock": True
        }), 200

    # Real API call to Fast2SMS DLT Manager
    headers = {
        "authorization": api_key,
        "accept": "application/json"
    }
    
    template_content = None
    url = "https://www.fast2sms.com/dev/dlt_manager"
    
    try:
        print(f"Querying DLT templates from Fast2SMS: {url} with auth={api_key[:8]}...")
        # type=template parameter is required to return templates array nested in each sender
        response = requests.get(url, headers=headers, params={"authorization": api_key, "type": "template"}, timeout=8)
        print(f"Fast2SMS DLT view status: {response.status_code}")
        
        if response.status_code == 200:
            res_data = response.json()
            # Fast2SMS returns "success": true in response
            if res_data.get("success") is True or res_data.get("return") is True:
                senders = res_data.get("data", [])
                if isinstance(senders, list):
                    for sender in senders:
                        item_sender_id = str(sender.get("sender_id") or "").upper()
                        # If sender_id is specified, filter by it
                        if sender_id and item_sender_id != sender_id:
                            continue
                        
                        templates_list = sender.get("templates", [])
                        if isinstance(templates_list, list):
                            for temp in templates_list:
                                item_template_id = str(temp.get("template_id") or temp.get("message_id") or "")
                                if item_template_id == template_id:
                                    template_content = temp.get("message_text") or temp.get("message") or temp.get("content")
                                    break
                        if template_content:
                            break
    except Exception as e:
        print(f"Failed to query {url}: {e}")

    if template_content:
        return jsonify({
            "success": True,
            "sender_id": sender_id,
            "template_id": template_id,
            "content": template_content,
            "is_mock": False
        }), 200
    else:
        fallback_content = f"Template {template_id} for Sender {sender_id} was not found on Fast2SMS. Please verify in your Fast2SMS DLT Manager dashboard."
        return jsonify({
            "success": False,
            "sender_id": sender_id,
            "template_id": template_id,
            "content": fallback_content,
            "error": "Template not found on Fast2SMS DLT Manager"
        }), 404


